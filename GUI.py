import tkinter as tk
from tkinter import filedialog, messagebox, simpledialog
import pandas as pd
from openpyxl import load_workbook

# Function to merge specific rows and columns from Excel 2 into Excel 1
def merge_sheets(file_path_1, file_path_2, output_path, start_row, start_col, row_range, col_range):
    # Load the workbooks and select the first sheets from both files
    workbook1 = load_workbook(file_path_1)
    workbook2 = load_workbook(file_path_2)
    sheet1 = workbook1.active  # First sheet in the first file
    sheet2 = workbook2.active  # First sheet in the second file

    # Parse the row and column ranges
    start_row_2, end_row_2 = map(int, row_range.split('-'))
    start_col_2, end_col_2 = map(int, col_range.split('-'))

    # Copy the data from sheet2 to sheet1 at the specified location
    for i, row in enumerate(sheet2.iter_rows(min_row=start_row_2, max_row=end_row_2, min_col=start_col_2, max_col=end_col_2), start=start_row):
        for j, cell in enumerate(row, start=start_col):
            if cell.value is not None:  # Only copy non-empty cells
                sheet1.cell(row=i, column=j).value = cell.value

    # Save the modified workbook to the specified output path
    workbook1.save(output_path)

# Function triggered by the button to select files, input ranges, and save the merged file
def select_files_and_merge():
    # Select the first Excel file
    file_path_1 = filedialog.askopenfilename(title="Alege fisierul in care copiezi informatii", filetypes=[("Excel files", "*.xlsx *.xls *.xlsm")])
    if not file_path_1:
        return  # If no file is selected, return

    # Select the second Excel file
    file_path_2 = filedialog.askopenfilename(title="Alege fisierul din care copiezi informatii", filetypes=[("Excel files", "*.xlsx *.xls *.xlsm")])
    if not file_path_2:
        return  # If no file is selected, return

    # Ask for row range in Excel 2 (e.g., "1-10")
    row_range = simpledialog.askstring("Input", "Alege randurile pentru copiat din Excel 2 (e.g., '1-10'):")
    if not row_range:
        return  # If no input is provided, return

    # Ask for column range in Excel 2 (e.g., "1-5")
    col_range = simpledialog.askstring("Input", "Alege coloanele pentru copiat din Excel 2 (e.g., '1-5'):")
    if not col_range:
        return  # If no input is provided, return

    # Ask for start row in Excel 1
    start_row = simpledialog.askinteger("Input", "Alege randul unde sa copiezi in Excel 1:")
    if not start_row:
        return  # If no input is provided, return

    # Ask for start column in Excel 1
    start_col = simpledialog.askinteger("Input", "Alege coloana unde sa copiezi in Excel 1:")
    if not start_col:
        return  # If no input is provided, return

    # Select the location to save the merged Excel file
    output_path = filedialog.asksaveasfilename(title="Salveaza noul fisier combinat", defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx *.xls *.xlsm")])
    if not output_path:
        return  # If no save location is selected, return

    try:
        # Perform the merging operation
        merge_sheets(file_path_1, file_path_2, output_path, start_row, start_col, row_range, col_range)
        messagebox.showinfo("Succes", f"Fisiere combinate si salvate in: {output_path}")
    except Exception as e:
        messagebox.showerror("Error", f"Eroare: {e}")

# Create the main application window
app = tk.Tk()
app.title("Excel Mixer")

# Create and place the button
button = tk.Button(app, text="Selecteaza 2 fisiere pentru combinat", command=select_files_and_merge)
button.pack(pady=20)

# Start the Tkinter event loop
app.minsize(400,400)
app.mainloop()
